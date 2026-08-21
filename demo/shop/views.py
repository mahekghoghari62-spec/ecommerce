from datetime import timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.conf import settings
from django.views.decorators.http import require_POST
import razorpay
from crud.models import Product as CrudProduct, Order as CrudOrder, Payment as CrudPayment, ProductReview
from crud.models import Category
from .forms import AddressForm, ProfileForm, CustomerRegistrationForm
from .models import Address, CustomerProfile, SiteSettings, SupplierProfile
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm
from accounts.forms import CustomerLoginForm


def get_razorpay_client():
    return razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))


CATEGORY_ICONS = {
    "electronics": "bi-laptop",
    "fashion": "bi-bag-fill",
    "home-kitchen": "bi-house-door-fill",
    "beauty": "bi-brush",
    "grocery": "bi-basket3-fill",
    "appliances": "bi-plug-fill",
    "toys-baby-products": "bi-balloon-fill",
    "sports-fitness": "bi-trophy-fill",
    "books-stationery": "bi-book-fill",
}


def get_nav_categories(category_mode=None):
    # When no explicit mode is passed in (most views just call this with no
    # args), fall back to the site-wide setting so the nav bar is consistent
    # everywhere — login, profile, address pages, etc. — not just on pages
    # that happen to compute a mode from products/supplier.
    if category_mode is None:
        category_mode = SiteSettings.get_solo().category_mode

    categories = []
    for category in Category.objects.filter(parent__isnull=True, is_active=True):
        if get_visible_products(category_mode).filter(category_id__in=category.get_descendant_ids()).exists():
            categories.append((category.slug, category.name, CATEGORY_ICONS.get(category.slug, "bi-grid")))
    return categories


def get_visible_products(category_mode=None):
    if category_mode is None:
        category_mode = SiteSettings.get_solo().category_mode

    products = CrudProduct.objects.filter(status="active")
    if category_mode == "single":
        clothes = Category.objects.filter(name="Clothes", parent__isnull=True).first()
        clothes_ids = clothes.get_descendant_ids() if clothes else []
        return products.filter(category_id__in=clothes_ids)
    return products


def get_request_category_mode(request):
    return SiteSettings.get_solo().category_mode


def _category_mode_from_products(products):
    """Nav bar mode ne ordered/viewed product(s) na supplier sathe match kare chhe.
    Badha products ek j supplier na hoy to tena mode ne vapare chhe;
    mixed/unknown hoy to 'multiple' (fuller nav) par fallback thay chhe."""
    suppliers = {p.supplier_id for p in products if p.supplier_id}
    if len(suppliers) == 1:
        supplier = next(p.supplier for p in products if p.supplier_id)
        return supplier.category_mode if supplier else "multiple"
    return "multiple"


def home(request):
    category_mode = get_request_category_mode(request)
    all_products = list(get_visible_products(category_mode).order_by("-created")[:12])
    return render(request, "shop/home.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": None,
        "products": all_products[:4],
        "more_products": all_products[4:],
        "cart_map": request.session.get("cart", {}),
    })


def category_detail(request, category_key):
    category = get_object_or_404(Category, slug=category_key, is_active=True)
    category_mode = get_request_category_mode(request)
    products = get_visible_products(category_mode).filter(
        status="active", category_id__in=category.get_descendant_ids()
    ).order_by("-created")
    return render(request, "shop/category.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": category.slug,
        "category_label": category.name,
        "products": products,
        "cart_map": request.session.get("cart", {}),
    })


def product_detail(request, product_id):
    category_mode = get_request_category_mode(request)
    product = get_object_or_404(
        get_visible_products(category_mode).filter(status="active"),
        id=product_id,
    )

    gallery_images = list(product.images.all())
    variants = list(product.variants.all())
    reviews = product.reviews.select_related("user").all()

    related_products = get_visible_products(category_mode).filter(
        status="active"
    ).exclude(id=product.id)
    if product.category_id:
        related_products = related_products.filter(category_id__in=product.category.get_descendant_ids())
    related_products = related_products.order_by("-created")[:4]

    return render(request, "shop/product_detail.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": product.category.slug if product.category_id else None,
        "product": product,
        "gallery_images": gallery_images,
        "variants": variants,
        "reviews": reviews,
        "related_products": related_products,
        "cart_map": request.session.get("cart", {}),
    })


@login_required(login_url='shop:login')
def profile_view(request):
    profile, _ = CustomerProfile.objects.get_or_create(user=request.user)

    if request.method == "POST":
        form = ProfileForm(request.POST, instance=profile, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect("shop:profile")
    else:
        form = ProfileForm(instance=profile, user=request.user)

    return render(request, "shop/profile.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
        "form": form,
    })


@login_required(login_url='shop:login')
def address_list(request):
    addresses = request.user.addresses.all()
    return render(request, "shop/address_list.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
        "addresses": addresses,
    })


@login_required(login_url='shop:login')
def address_add(request):
    if request.method == "POST":
        form = AddressForm(request.POST)
        if form.is_valid():
            address = form.save(commit=False)
            address.user = request.user
            if address.is_default:
                request.user.addresses.update(is_default=False)
            address.save()
            messages.success(request, "Address added.")
            return redirect("shop:address_list")
    else:
        # Prefill full name and phone from the customer's saved profile so
        # they don't have to retype details already on file.
        profile = CustomerProfile.objects.filter(user=request.user).first()
        last_address = request.user.addresses.first()
        initial = {
            "full_name": request.user.get_full_name() or request.user.username,
            "phone": profile.phone if profile else "",
        }
        if last_address:
            initial["city"] = last_address.city
            initial["state"] = last_address.state
        form = AddressForm(initial=initial)
    return render(request, "shop/address_form.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
        "form": form,
        "title": "Add Address",
    })


@login_required(login_url='shop:login')
def address_edit(request, pk):
    address = get_object_or_404(Address, pk=pk, user=request.user)
    if request.method == "POST":
        form = AddressForm(request.POST, instance=address)
        if form.is_valid():
            updated = form.save(commit=False)
            if updated.is_default:
                request.user.addresses.exclude(pk=address.pk).update(is_default=False)
            updated.save()
            messages.success(request, "Address updated.")
            return redirect("shop:address_list")
    else:
        form = AddressForm(instance=address)
    return render(request, "shop/address_form.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
        "form": form,
        "title": "Edit Address",
    })


@login_required(login_url='shop:login')
def address_delete(request, pk):
    address = get_object_or_404(Address, pk=pk, user=request.user)
    if request.method == "POST":
        address.delete()
        messages.success(request, "Address deleted.")
        return redirect("shop:address_list")
    return render(request, "shop/address_confirm_delete.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
        "address": address,
    })


def shop_login(request):
    next_url = request.GET.get("next") or "shop:home"

    if request.user.is_authenticated:
        return redirect(next_url)

    if request.method == "POST":
        form = CustomerLoginForm(request, data=request.POST)
        if form.is_valid():
            auth_login(request, form.get_user())
            return redirect(request.POST.get("next") or next_url)
    else:
        form = CustomerLoginForm()

    return render(request, "shop/login.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
        "form": form,
        "next": next_url,
    })


def shop_register(request):
    next_url = request.GET.get("next") or "shop:home"

    if request.user.is_authenticated:
        return redirect(next_url)

    if request.method == "POST":
        form = CustomerRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user, backend="django.contrib.auth.backends.ModelBackend")
            return redirect(request.POST.get("next") or next_url)
    else:
        form = CustomerRegistrationForm()

    return render(request, "shop/register.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
        "form": form,
        "next": next_url,
    })


# ---------------------------
# Cart views
# ---------------------------

def _get_price(product):
    """Always use the GST-inclusive selling price where available, so the
    amount shown here matches the product detail page (product.selling_price
    = price + gst_percent). Falls back to plain .price for models without
    a selling_price property."""
    selling_price = getattr(product, "selling_price", None)
    if selling_price:
        return selling_price
    return getattr(product, "price", 0)


def cart_view(request):
    cart = request.session.get("cart", {})
    active_items = []
    subtotal = 0
    total_items = 0

    for product_id, qty in cart.items():
        product = CrudProduct.objects.filter(id=product_id).first()
        if product:
            price = _get_price(product)
            line_total = price * qty
            subtotal += line_total
            total_items += qty
            active_items.append({
                "id": product_id,
                "product": product,
                "quantity": qty,
                "line_total": line_total,
            })

    category_mode = _category_mode_from_products([item["product"] for item in active_items])

    saved = request.session.get("saved_for_later", {})
    saved_items = []
    for product_id in saved:
        product = CrudProduct.objects.filter(id=product_id).first()
        if product:
            saved_items.append({"id": product_id, "product": product})

    related_products = get_visible_products(category_mode).exclude(
        id__in=list(cart.keys()) + list(saved.keys())
    ).order_by("-created")[:6]

    return render(request, "shop/cart.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": None,
        "active_items": active_items,
        "saved_items": saved_items,
        "related_products": related_products,
        "total_items": total_items,
        "subtotal": subtotal,
        "delivery_date": timezone.now() + timedelta(days=5),
    })


def _is_ajax(request):
    return request.headers.get("x-requested-with") == "XMLHttpRequest"


def add_to_cart(request, product_id):
    product = get_object_or_404(CrudProduct, id=product_id)
    cart = request.session.get("cart", {})
    cart[str(product_id)] = cart.get(str(product_id), 0) + 1
    request.session["cart"] = cart
    request.session.modified = True

    if _is_ajax(request):
        return JsonResponse({
            "success": True,
            "message": f"{product.name} added to cart.",
            "quantity": cart[str(product_id)],
            "cart_count": sum(cart.values()),
        })

    messages.success(request, f"{product.name} added to cart.")
    next_url = request.POST.get("next")
    return redirect(next_url or "shop:cart")


def update_cart_item(request, item_id):
    cart = request.session.get("cart", {})
    item_id = str(item_id)
    action = request.POST.get("action")
    removed = False
    qty = cart.get(item_id, 0)

    if item_id in cart:
        if action == "increase":
            cart[item_id] += 1
            qty = cart[item_id]
        elif action == "decrease":
            cart[item_id] -= 1
            if cart[item_id] <= 0:
                cart.pop(item_id, None)
                qty = 0
                removed = True
            else:
                qty = cart[item_id]
        request.session["cart"] = cart
        request.session.modified = True

    if _is_ajax(request):
        return JsonResponse({
            "success": True,
            "quantity": qty,
            "removed": removed,
            "cart_count": sum(cart.values()),
        })

    return redirect("shop:cart")


def remove_cart_item(request, item_id):
    cart = request.session.get("cart", {})
    cart.pop(str(item_id), None)
    request.session["cart"] = cart
    messages.success(request, "Item removed from cart.")
    return redirect("shop:cart")


def save_for_later(request, item_id):
    cart = request.session.get("cart", {})
    saved = request.session.get("saved_for_later", {})
    if str(item_id) in cart:
        saved[str(item_id)] = cart.pop(str(item_id))
        request.session["cart"] = cart
        request.session["saved_for_later"] = saved
    return redirect("shop:cart")


def move_to_cart(request, item_id):
    cart = request.session.get("cart", {})
    saved = request.session.get("saved_for_later", {})
    if str(item_id) in saved:
        cart[str(item_id)] = saved.pop(str(item_id))
        request.session["cart"] = cart
        request.session["saved_for_later"] = saved
    return redirect("shop:cart")


# ---------------------------
# Checkout / Place Order / Razorpay
# ---------------------------

@login_required(login_url='shop:login')
def place_order_view(request):
    cart = request.session.get("cart", {})
    if not cart:
        messages.warning(request, "Your cart is empty.")
        return redirect("shop:cart")

    addresses = request.user.addresses.all()
    active_items = []
    subtotal = 0

    for product_id, qty in cart.items():
        product = CrudProduct.objects.filter(id=product_id).first()
        if product:
            price = _get_price(product)
            line_total = price * qty
            subtotal += line_total
            active_items.append({
                "product": product,
                "quantity": qty,
                "line_total": line_total,
            })

    category_mode = _category_mode_from_products([item["product"] for item in active_items])

    PAYMENT_METHOD_MAP = {
        "cod": "cash",
        "upi": "upi",
        "card": "card",
    }

    if request.method == "POST":
        address_id = request.POST.get("address_id")
        payment_choice = request.POST.get("payment_method")

        if not address_id:
            messages.error(request, "Please select a delivery address.")
            return redirect("shop:place_order")

        if payment_choice not in PAYMENT_METHOD_MAP:
            messages.error(request, "Please select a payment method.")
            return redirect("shop:place_order")

        address = get_object_or_404(Address, pk=address_id, user=request.user)
        payment_method = PAYMENT_METHOD_MAP[payment_choice]

        razorpay_payment_id = request.POST.get("razorpay_payment_id", "")
        razorpay_order_id = request.POST.get("razorpay_order_id", "")
        razorpay_signature = request.POST.get("razorpay_signature", "")

        if payment_choice in ("upi", "card"):
            if not (razorpay_payment_id and razorpay_order_id and razorpay_signature):
                messages.error(request, "Payment verification failed. Please try again.")
                return redirect("shop:place_order")

            client = get_razorpay_client()
            try:
                client.utility.verify_payment_signature({
                    "razorpay_order_id": razorpay_order_id,
                    "razorpay_payment_id": razorpay_payment_id,
                    "razorpay_signature": razorpay_signature,
                })
            except razorpay.errors.SignatureVerificationError:
                messages.error(request, "Payment verification failed. Please try again.")
                return redirect("shop:place_order")

            payment_status = "completed"
            transaction_id = razorpay_payment_id
        else:
            payment_status = "pending"
            transaction_id = ""

        created_order_ids = []

        for product_id, qty in cart.items():
            product = CrudProduct.objects.filter(id=product_id).first()
            if not product:
                continue
            price = _get_price(product)
            line_amount = price * qty

            order = CrudOrder.objects.create(
                customer_name=request.user.get_full_name() or request.user.username,
                customer_email=request.user.email,
                product=product,
                quantity=qty,
                amount=line_amount,
                status="pending",
            )
            created_order_ids.append(order.id)

            CrudPayment.objects.create(
                order=order,
                amount=line_amount,
                method=payment_method,
                status=payment_status,
                transaction_id=transaction_id,
                notes=f"Payment via {payment_choice.upper()} (checkout)",
            )

        request.session["cart"] = {}
        # Store this order batch + delivery address snapshot so the
        # invoice/success page (and its PDF/Excel downloads) can render it.
        request.session["last_order_ids"] = created_order_ids
        request.session["last_order_address"] = {
            "full_name": address.full_name,
            "phone": address.phone,
            "address_line1": address.address_line1,
            "address_line2": address.address_line2,
            "city": address.city,
            "state": address.state,
            "pincode": address.pincode,
        }
        request.session["last_order_payment_choice"] = payment_choice
        request.session.modified = True

        messages.success(request, "Your order has been placed successfully!")
        return redirect("shop:order_success")

    return render(request, "shop/place_order.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": None,
        "addresses": addresses,
        "active_items": active_items,
        "subtotal": subtotal,
    })


@login_required(login_url='shop:login')
@require_POST
def create_razorpay_order(request):
    cart = request.session.get("cart", {})
    if not cart:
        return JsonResponse({"success": False, "error": "Cart is empty."})

    subtotal = 0
    for product_id, qty in cart.items():
        product = CrudProduct.objects.filter(id=product_id).first()
        if product:
            price = _get_price(product)
            subtotal += price * qty

    amount_paise = int(subtotal * 100)

    client = get_razorpay_client()
    razorpay_order = client.order.create({
        "amount": amount_paise,
        "currency": "INR",
        "payment_capture": 1,
    })

    return JsonResponse({
        "success": True,
        "order_id": razorpay_order["id"],
        "amount": amount_paise,
        "key": settings.RAZORPAY_KEY_ID,
        "name": request.user.get_full_name() or request.user.username,
        "email": request.user.email,
    })


@login_required(login_url='shop:login')
def order_success_view(request):
    order_ids = request.session.get("last_order_ids", [])
    orders = CrudOrder.objects.filter(
        id__in=order_ids, customer_email=request.user.email
    ).select_related("product", "product__supplier")

    ordered_products = [o.product for o in orders if o.product_id]
    category_mode = _category_mode_from_products(ordered_products)

    invoice_items = []
    grand_total = 0
    for order in orders:
        invoice_items.append({
            "order_number": order.order_number,
            "product_name": order.product.name,
            "quantity": order.quantity,
            "unit_price": (order.amount / order.quantity) if order.quantity else 0,
            "line_total": order.amount,
        })
        grand_total += order.amount

    address = request.session.get("last_order_address", {})
    payment_choice = request.session.get("last_order_payment_choice", "")
    payment_labels = {"cod": "Cash on Delivery", "upi": "UPI", "card": "Credit / Debit Card"}

    return render(request, "shop/order_success.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": None,
        "invoice_items": invoice_items,
        "grand_total": grand_total,
        "address": address,
        "payment_label": payment_labels.get(payment_choice, payment_choice),
        "has_invoice": bool(invoice_items),
        "invoice_number": invoice_items[0]["order_number"] if invoice_items else "",
    })


def _get_invoice_context(request):
    """Shared helper: rebuild the invoice line items + totals from session
    order ids, scoped to the logged-in user's own orders."""
    order_ids = request.session.get("last_order_ids", [])
    orders = CrudOrder.objects.filter(id__in=order_ids, customer_email=request.user.email).select_related("product")

    items = []
    grand_total = 0
    order_date = None
    for order in orders:
        if order_date is None or order.order_date < order_date:
            order_date = order.order_date
        items.append({
            "order_number": order.order_number,
            "product_name": order.product.name,
            "quantity": order.quantity,
            "unit_price": (order.amount / order.quantity) if order.quantity else 0,
            "line_total": order.amount,
        })
        grand_total += order.amount

    address = request.session.get("last_order_address", {})
    return items, grand_total, address, order_date


@login_required(login_url='shop:login')
def invoice_pdf(request):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

    items, grand_total, address, order_date = _get_invoice_context(request)
    if not items:
        messages.error(request, "No recent order found to generate an invoice.")
        return redirect("shop:order_success")

    invoice_no = items[0]["order_number"]
    invoice_date = (order_date or timezone.now()).strftime("%d %b %Y")

    PINK = colors.HexColor("#9F2089")
    LIGHT_PINK = colors.HexColor("#FCE9F9")
    GREY = colors.HexColor("#666666")

    styles = getSampleStyleSheet()
    # NOTE: fontSize alone does not change a ParagraphStyle's line height —
    # `leading` must be set explicitly (rule of thumb: fontSize * 1.2),
    # otherwise a bigger font renders taller than its inherited leading and
    # overlaps the paragraph below it. Both styles below previously omitted
    # `leading`, which caused "MyShop" / "TAX INVOICE" to overlap the text
    # right under them.
    company_style = ParagraphStyle("company", parent=styles["Normal"], fontSize=18, leading=22, textColor=PINK, fontName="Helvetica-Bold")
    tagline_style = ParagraphStyle("tagline", parent=styles["Normal"], fontSize=8.5, textColor=GREY)
    meta_label_style = ParagraphStyle("metaLabel", parent=styles["Normal"], fontSize=9, textColor=GREY, alignment=TA_RIGHT)
    meta_value_style = ParagraphStyle("metaValue", parent=styles["Normal"], fontSize=10, textColor=colors.black, alignment=TA_RIGHT, fontName="Helvetica-Bold")
    title_style = ParagraphStyle("title", parent=styles["Normal"], fontSize=13, leading=16, alignment=TA_RIGHT, fontName="Helvetica-Bold")
    section_label_style = ParagraphStyle("sectionLabel", parent=styles["Normal"], fontSize=9, textColor=PINK, fontName="Helvetica-Bold")
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=9.5, leading=14)
    footer_style = ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=GREY, alignment=TA_LEFT)
    thanks_style = ParagraphStyle("thanks", parent=styles["Normal"], fontSize=10, leading=13, textColor=PINK, fontName="Helvetica-Bold")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=16 * mm, bottomMargin=16 * mm,
        leftMargin=16 * mm, rightMargin=16 * mm,
    )
    elements = []

    # --- Header: company block (left) + invoice meta (right) ---
    company_block = [
        Paragraph("MyShop", company_style),
        Paragraph("Meesho-style Marketplace Pvt. Ltd.", tagline_style),
        Paragraph("www.myshop.example &nbsp;|&nbsp; support@myshop.example", tagline_style),
    ]
    meta_block = [
        Paragraph("TAX INVOICE", title_style),
        Spacer(1, 4),
        Paragraph(f"Invoice No: <b>{invoice_no}</b>", meta_label_style),
        Paragraph(f"Invoice Date: <b>{invoice_date}</b>", meta_label_style),
    ]
    header_table = Table([[company_block, meta_block]], colWidths=[100 * mm, 74 * mm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=1.2, color=PINK))
    elements.append(Spacer(1, 12))

    # --- Bill To box ---
    if address:
        addr_line = ", ".join(filter(None, [
            address.get("address_line1", ""), address.get("address_line2", ""),
        ]))
        city_line = ", ".join(filter(None, [
            address.get("city", ""), address.get("state", ""), address.get("pincode", ""),
        ]))
        bill_to_content = [
            Paragraph("BILL TO / SHIP TO", section_label_style),
            Spacer(1, 4),
            Paragraph(f"<b>{address.get('full_name', '')}</b>", body_style),
            Paragraph(addr_line, body_style),
            Paragraph(city_line, body_style),
            Paragraph(f"Phone: {address.get('phone', '')}", body_style),
        ]
        bill_to_table = Table([[bill_to_content]], colWidths=[174 * mm])
        bill_to_table.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#dddddd")),
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_PINK),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(bill_to_table)
        elements.append(Spacer(1, 16))

    # --- Items table ---
    table_data = [["#", "Product", "Qty", "Unit Price", "Amount"]]
    for idx, item in enumerate(items, start=1):
        table_data.append([
            str(idx),
            item["product_name"],
            str(item["quantity"]),
            f"Rs. {item['unit_price']:.2f}",
            f"Rs. {item['line_total']:.2f}",
        ])

    items_table = Table(table_data, colWidths=[12 * mm, 78 * mm, 20 * mm, 32 * mm, 32 * mm], repeatRows=1)
    row_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), PINK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, PINK),
        ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#e5e5e5")),
    ]
    for r in range(1, len(table_data)):
        if r % 2 == 0:
            row_styles.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#faf5fa")))
    items_table.setStyle(TableStyle(row_styles))
    elements.append(items_table)
    elements.append(Spacer(1, 10))

    # --- Totals box (right aligned) ---
    totals_data = [
        ["Subtotal", f"Rs. {grand_total:.2f}"],
        ["Delivery Charges", "FREE"],
        ["Grand Total", f"Rs. {grand_total:.2f}"],
    ]
    totals_table = Table(totals_data, colWidths=[35 * mm, 32 * mm])
    totals_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEABOVE", (0, -1), (-1, -1), 1, PINK),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 11),
        ("TEXTCOLOR", (0, -1), (-1, -1), PINK),
    ]))
    wrapper = Table([["", totals_table]], colWidths=[107 * mm, 67 * mm])
    elements.append(wrapper)
    elements.append(Spacer(1, 22))

    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#dddddd")))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        "This is a computer-generated invoice and does not require a physical signature. "
        "For any queries regarding this order, please contact support@myshop.example.",
        footer_style,
    ))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph("Thank you for shopping with MyShop!", thanks_style))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice_{invoice_no}.pdf"'
    return response


@login_required(login_url='shop:login')
def invoice_excel(request):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items, grand_total, address, order_date = _get_invoice_context(request)
    if not items:
        messages.error(request, "No recent order found to generate an invoice.")
        return redirect("shop:order_success")

    invoice_no = items[0]["order_number"]
    invoice_date = (order_date or timezone.now()).strftime("%d %b %Y")

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    PINK = "9F2089"
    LIGHT_PINK = "FCE9F9"
    header_fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
    light_fill = PatternFill(start_color=LIGHT_PINK, end_color=LIGHT_PINK, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    grey_font = Font(color="777777", size=9)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"),
    )

    ws.merge_cells("A1:E1")
    ws["A1"] = "MyShop"
    ws["A1"].font = Font(size=16, bold=True, color=PINK)

    ws.merge_cells("A2:E2")
    ws["A2"] = "Meesho-style Marketplace Pvt. Ltd. | support@myshop.example"
    ws["A2"].font = grey_font

    ws.merge_cells("A4:E4")
    ws["A4"] = "TAX INVOICE"
    ws["A4"].font = Font(size=13, bold=True)

    ws["A6"] = "Invoice No:"
    ws["A6"].font = bold_font
    ws["B6"] = invoice_no
    ws["D6"] = "Invoice Date:"
    ws["D6"].font = bold_font
    ws["E6"] = invoice_date

    if address:
        ws.merge_cells("A8:E8")
        ws["A8"] = "BILL TO / SHIP TO"
        ws["A8"].font = Font(bold=True, color=PINK)

        ws["A9"] = address.get("full_name", "")
        ws["A9"].font = bold_font
        addr_line = ", ".join(filter(None, [
            address.get("address_line1", ""), address.get("address_line2", ""),
        ]))
        city_line = ", ".join(filter(None, [
            address.get("city", ""), address.get("state", ""), address.get("pincode", ""),
        ]))
        ws["A10"] = addr_line
        ws["A11"] = city_line
        ws["A12"] = f"Phone: {address.get('phone', '')}"
        for row in (9, 10, 11, 12):
            ws.merge_cells(f"A{row}:E{row}")

    start_row = 14
    headers = ["#", "Product", "Qty", "Unit Price (Rs.)", "Amount (Rs.)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row = start_row + 1
    for idx, item in enumerate(items, start=1):
        values = [idx, item["product_name"], item["quantity"], round(float(item["unit_price"]), 2), round(float(item["line_total"]), 2)]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in (1, 3, 4, 5):
                cell.alignment = Alignment(horizontal="center" if col in (1, 3) else "right")
            if idx % 2 == 0:
                cell.fill = light_fill
        row += 1

    ws.cell(row=row, column=4, value="Grand Total").font = bold_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=row, column=5, value=round(float(grand_total), 2))
    total_cell.font = Font(bold=True, color=PINK, size=12)
    total_cell.alignment = Alignment(horizontal="right")

    for col, width in zip("ABCDE", [6, 34, 8, 18, 18]):
        ws.column_dimensions[col].width = width

    footer_row = row + 3
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    ws[f"A{footer_row}"] = "This is a computer-generated invoice. Thank you for shopping with MyShop!"
    ws[f"A{footer_row}"].font = grey_font

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="invoice_{invoice_no}.xlsx"'
    return response


@login_required(login_url='shop:login')
@require_POST
def add_review(request, product_id):
    product = get_object_or_404(CrudProduct, id=product_id)
    rating = request.POST.get("rating")
    comment = request.POST.get("comment", "").strip()

    if not rating or not rating.isdigit() or not (1 <= int(rating) <= 5):
        messages.error(request, "Please select a rating between 1 and 5.")
        return redirect("shop:product_detail", product_id=product_id)

    # One review per user per product — resubmitting updates their existing review.
    ProductReview.objects.update_or_create(
        product=product,
        user=request.user,
        defaults={"rating": int(rating), "comment": comment},
    )
    messages.success(request, "Thanks for your review!")
    return redirect("shop:product_detail", product_id=product_id)


def shop_logout(request):
    auth_logout(request)
    return redirect("shop:home")
from crud.views import LEGAL_POLICIES, LEGAL_POLICY_BODY


def policy_detail(request, slug):
    policy = next((p for p in LEGAL_POLICIES if p["slug"] == slug), None)
    category_mode = get_request_category_mode(request)
    return render(request, "shop/policy_detail.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": None,
        "policy": policy,
        "policy_body": LEGAL_POLICY_BODY.get(slug, []),
    })
COMPANY_INFO_PAGES = {
    "about-us": {
        "title": "About Us",
        "body": [
            "MyShop is an online marketplace connecting suppliers and customers across India, offering a wide range of products at affordable prices.",
            "Founded with a mission to make quality products accessible to everyone, we work with thousands of suppliers nationwide.",
        ],
    },
    "careers": {
        "title": "Careers",
        "body": [
            "We're always looking for passionate people to join our team.",
            "Currently there are no open positions listed. Please check back soon, or write to us at careers@myshop.example.",
        ],
    },
    "stories": {
        "title": "MyShop Stories",
        "body": [
            "Read about the suppliers, customers, and communities that make MyShop possible.",
            "New stories are published regularly — check back soon for the latest features.",
        ],
    },
    "press": {
        "title": "Press",
        "body": [
            "For press inquiries, media kits, or interview requests, please contact press@myshop.example.",
        ],
    },
    "corporate-information": {
        "title": "Corporate Information",
        "body": [
            "MyShop Marketplace Pvt. Limited",
            "CIN: U51109GJ2024PTC000000",
            "Registered Office: Business Bay, Sector 21, Ahmedabad, 380015, Gujarat, India",
        ],
    },
    "advertise": {
        "title": "Advertise With Us",
        "body": [
            "Grow your brand's visibility with MyShop advertising solutions.",
            "For advertising partnerships, write to us at advertise@myshop.example.",
        ],
    },
    "gift-cards": {
        "title": "Gift Cards",
        "body": [
            "MyShop gift cards are coming soon! Stay tuned for updates.",
        ],
    },
    "help-center": {
        "title": "Help Center",
        "body": [
            "Need assistance? Reach out to our support team at support@myshop.example and we'll get back to you within 24 hours.",
        ],
    },
    "sitemap": {
        "title": "Sitemap",
        "body": [
            "Browse MyShop by category using the navigation bar at the top of every page, or visit the homepage to explore all products.",
        ],
    },
}


def info_page(request, slug):
    page = COMPANY_INFO_PAGES.get(slug)
    category_mode = get_request_category_mode(request)
    return render(request, "shop/info_page.html", {
        "nav_categories": get_nav_categories(category_mode),
        "current_category": None,
        "page": page,
    })
# ---------------------------------------------------------------------------
# Footer pages — each has its own template (per-page, not a generic slug page)
# Paste these into shop/views.py, e.g. right after `home()` or at the bottom.
# They all share the same nav_categories context as the rest of the site.
# ---------------------------------------------------------------------------

def about_us(request):
    return render(request, "shop/about_us.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def contact_us(request):
    return render(request, "shop/contact_us.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def careers(request):
    return render(request, "shop/careers.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def myshop_stories(request):
    return render(request, "shop/myshop_stories.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def press(request):
    return render(request, "shop/press.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def corporate_information(request):
    return render(request, "shop/corporate_information.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def payments(request):
    return render(request, "shop/payments.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def faq(request):
    return render(request, "shop/faq.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def sitemap(request):
    return render(request, "shop/sitemap.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def grievance_redressal(request):
    return render(request, "shop/grievance_redressal.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def advertise(request):
    return render(request, "shop/advertise.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def gift_cards(request):
    return render(request, "shop/gift_cards.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })


def help_center(request):
    return render(request, "shop/help_center.html", {
        "nav_categories": get_nav_categories(),
        "current_category": None,
    })
