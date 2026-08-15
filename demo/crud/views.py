from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.views.generic import CreateView, DeleteView, DetailView, TemplateView,UpdateView
from django.db import models
from django_filters.views import FilterView
from django_tables2 import SingleTableMixin
from .mixins import AjaxModalDeleteMixin, AjaxModalFormMixin
from django.views.generic import TemplateView
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import redirect
from .filters import AdvertisementFilter, CatalogUploadFilter, ClaimFilter, ContactFilter, ImageBulkUploadFilter, InfluencerCampaignFilter, InventoryFilter, OrderFilter, PaymentFilter, PricingFilter, ProductFilter, ProjectFilter, PromotionFilter, QualityFilter, ReturnFilter, WarehouseFilter
from .forms import AdvertisementForm, CallbackRequestForm, CatalogUploadForm, ClaimForm, ContactForm, ImageBulkUploadForm, InfluencerCampaignForm, InventoryForm, OrderForm, PaymentForm, PricingForm, ProductForm, PromotionForm, QualityForm, ReturnForm, WarehouseForm
from .models import Advertisement, CallbackRequest, CatalogUpload, Claim, Contact, DailyMetric, ImageBulkUpload, InfluencerCampaign, Inventory, Order, Payment, Pricing, Product, Project, Promotion, Quality, Return,SupportTicket, Warehouse, PayoutCycle, CompensationRecovery
from .tables import AdvertisementTable, CatalogUploadTable, ClaimTable, ContactTable, ImageBulkUploadTable, InfluencerCampaignTable, InventoryTable, OrderTable, PaymentTable, PricingTable, ProductTable, ProjectTable, PromotionTable, QualityTable, ReturnTable, WarehouseTable
import csv
import io
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.views import View
from .forms import InventoryBulkStockUploadForm
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.urls import reverse
from shop.models import SupplierProfile
from .forms import (
    WhatsAppSettingsForm, BankDetailsForm, TaxDetailsForm,
    SupplierSignatureForm, EmailNotificationsForm,
)
# --- Orders: full CRUD (tables2 + django-filter + crispy form + messages) ---
class OrderListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Order
    table_class = OrderTable
    filterset_class = OrderFilter
    template_name = "crud/order_list.html"
    table_pagination = {"per_page": 10}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["status_tabs"] = [
            {"key": key, "label": label, "count": Order.objects.filter(status=key).count()}
            for key, label in Order.STATUS_CHOICES
        ]
        context["total_count"] = Order.objects.count()
        context["active_status"] = self.request.GET.get("status", "")
        return context


class OrderCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Order
    form_class = OrderForm
    template_name = "crud/order_form.html"
    success_url = reverse_lazy("crud:order_list")

    def form_valid(self, form):
        messages.success(self.request, f"Order created for “{form.instance.customer_name}”.")
        return super().form_valid(form)


class OrderUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Order
    form_class = OrderForm
    template_name = "crud/order_form.html"
    success_url = reverse_lazy("crud:order_list")

    def form_valid(self, form):
        messages.success(self.request, f"Order “{form.instance.order_number}” updated.")
        return super().form_valid(form)


class OrderDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Order
    template_name = "crud/order_confirm_delete.html"
    success_url = reverse_lazy("crud:order_list")

    def form_valid(self, form):
        messages.success(self.request, f"Order “{self.object.order_number}” deleted.")
        return super().form_valid(form)


# --- Contacts: full CRUD (tables2 + django-filter + crispy form + messages) ---
class ContactListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Contact
    table_class = ContactTable
    filterset_class = ContactFilter
    template_name = "crud/contact_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return super().get_queryset().select_related("company")


class ContactCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Contact
    form_class = ContactForm
    template_name = "crud/contact_form.html"
    success_url = reverse_lazy("crud:contact_list")

    def form_valid(self, form):
        messages.success(self.request, f"Contact “{form.instance.name}” created.")
        return super().form_valid(form)


class ContactUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Contact
    form_class = ContactForm
    template_name = "crud/contact_form.html"
    success_url = reverse_lazy("crud:contact_list")

    def form_valid(self, form):
        messages.success(self.request, f"Contact “{form.instance.name}” updated.")
        return super().form_valid(form)


class ContactDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Contact
    template_name = "crud/contact_confirm_delete.html"
    success_url = reverse_lazy("crud:contact_list")

    def form_valid(self, form):
        messages.success(self.request, f"Contact “{self.object.name}” deleted.")
        return super().form_valid(form)


# --- Projects: relational list + detail (showcases FK / M2M data) ---
class ProjectListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Project
    table_class = ProjectTable
    filterset_class = ProjectFilter
    template_name = "crud/project_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return (
            super().get_queryset()
            .select_related("company", "lead")
            .prefetch_related("team", "tags")
        )


class ProjectDetailView(LoginRequiredMixin, DetailView):
    model = Project
    template_name = "crud/project_detail.html"

    def get_queryset(self):
        return (
            super().get_queryset()
            .select_related("company", "lead")
            .prefetch_related("team", "tags", "tasks__assignee")
        )


# --- Returns: full CRUD (tables2 + django-filter + crispy form + messages) ---
class ReturnListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Return
    table_class = ReturnTable
    filterset_class = ReturnFilter
    template_name = "crud/return_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        qs = super().get_queryset().select_related("order")
        tab = self.request.GET.get("tab", "")
        if tab in ("requested", "approved", "picked_up", "refunded", "rejected"):
            qs = qs.filter(status=tab)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_returns = Return.objects.all()
        context["total_returns"] = all_returns.count()
        context["requested_count"] = all_returns.filter(status="requested").count()
        context["approved_count"] = all_returns.filter(status="approved").count()
        context["rejected_count"] = all_returns.filter(status="rejected").count()
        context["picked_up_count"] = all_returns.filter(status="picked_up").count()
        context["refunded_count"] = all_returns.filter(status="refunded").count()

        tab_param = self.request.GET.get("tab", "")
        context["show_claim_tracking"] = tab_param == "claims"
        context["show_return_tracking"] = tab_param != "" and tab_param != "claims"
        context["active_tab"] = tab_param

        all_claims = Claim.objects.select_related("order")
        claim_tab = self.request.GET.get("claim_tab", "all")
        context["active_claim_tab"] = claim_tab
        filtered_claims = all_claims
        if claim_tab in ("open", "under_review", "approved", "rejected", "settled"):
            filtered_claims = all_claims.filter(status=claim_tab)
        context["claim_table"] = ClaimTable(filtered_claims)
        context["total_claims"] = all_claims.count()
        context["open_claims_count"] = all_claims.filter(status="open").count()
        context["under_review_claims_count"] = all_claims.filter(status="under_review").count()
        context["approved_claims_count"] = all_claims.filter(status="approved").count()
        context["rejected_claims_count"] = all_claims.filter(status="rejected").count()
        context["settled_claims_count"] = all_claims.filter(status="settled").count()
        return context

class ReturnCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Return
    form_class = ReturnForm
    template_name = "crud/return_form.html"
    success_url = reverse_lazy("crud:return_list")

    def form_valid(self, form):
        messages.success(self.request, f"Return created for order “{form.instance.order.order_number}”.")
        return super().form_valid(form)


class ReturnUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Return
    form_class = ReturnForm
    template_name = "crud/return_form.html"
    success_url = reverse_lazy("crud:return_list")

    def form_valid(self, form):
        messages.success(self.request, f"Return for “{form.instance.order.order_number}” updated.")
        return super().form_valid(form)


class ReturnDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Return
    template_name = "crud/return_confirm_delete.html"
    success_url = reverse_lazy("crud:return_list")

    def form_valid(self, form):
        messages.success(self.request, f"Return for “{self.object.order.order_number}” deleted.")
        return super().form_valid(form)


# --- Pricing: full CRUD (tables2 + django-filter + crispy form + messages) ---
class PricingListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Pricing
    table_class = PricingTable
    filterset_class = PricingFilter
    template_name = "crud/pricing_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return super().get_queryset().select_related("product")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_pricing = Pricing.objects.select_related("product")
        context["total_pricing"] = all_pricing.count()
        context["active_count"] = all_pricing.filter(status="active").count()
        context["draft_count"] = all_pricing.filter(status="draft").count()
        context["archived_count"] = all_pricing.filter(status="archived").count()

        context["active_table"] = PricingTable(all_pricing.filter(status="active"), prefix="active-")
        context["draft_table"] = PricingTable(all_pricing.filter(status="draft"), prefix="draft-")
        context["archived_table"] = PricingTable(all_pricing.filter(status="archived"), prefix="archived-")

        active_items = all_pricing.filter(status="active")
        if active_items.exists():
            avg_margin = sum(p.margin_percent for p in active_items) / active_items.count()
            context["avg_margin"] = round(avg_margin, 1)
        else:
            context["avg_margin"] = 0

        context["losing_margin_count"] = active_items.filter(
            selling_price__lte=models.F("cost_price")
        ).count()
        return context
class PricingCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Pricing
    form_class = PricingForm
    template_name = "crud/pricing_form.html"
    success_url = reverse_lazy("crud:pricing_list")

    def form_valid(self, form):
        messages.success(self.request, f"Pricing created for “{form.instance.product.name}”.")
        return super().form_valid(form)


class PricingUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Pricing
    form_class = PricingForm
    template_name = "crud/pricing_form.html"
    success_url = reverse_lazy("crud:pricing_list")

    def form_valid(self, form):
        messages.success(self.request, f"Pricing for “{form.instance.product.name}” updated.")
        return super().form_valid(form)


class PricingDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Pricing
    template_name = "crud/pricing_confirm_delete.html"
    success_url = reverse_lazy("crud:pricing_list")

    def form_valid(self, form):
        messages.success(self.request, f"Pricing for “{self.object.product.name}” deleted.")
        return super().form_valid(form)


# --- Claims: full CRUD (tables2 + django-filter + crispy form + messages) ---
class ClaimListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Claim
    table_class = ClaimTable
    filterset_class = ClaimFilter
    template_name = "crud/claim_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return super().get_queryset().select_related("order")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_claims = Claim.objects.all()
        context["total_claims"] = all_claims.count()
        context["open_count"] = all_claims.filter(status="open").count()
        context["approved_count"] = all_claims.filter(status="approved").count()
        context["settled_count"] = all_claims.filter(status="settled").count()
        return context

class ClaimCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Claim
    form_class = ClaimForm
    template_name = "crud/claim_form.html"
    success_url = reverse_lazy("crud:claim_list")

    def form_valid(self, form):
        messages.success(self.request, f"Claim created for order “{form.instance.order.order_number}”.")
        return super().form_valid(form)


class ClaimUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Claim
    form_class = ClaimForm
    template_name = "crud/claim_form.html"
    success_url = reverse_lazy("crud:claim_list")

    def form_valid(self, form):
        messages.success(self.request, f"Claim for “{form.instance.order.order_number}” updated.")
        return super().form_valid(form)


class ClaimDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Claim
    template_name = "crud/claim_confirm_delete.html"
    success_url = reverse_lazy("crud:claim_list")

    def form_valid(self, form):
        messages.success(self.request, f"Claim for “{self.object.order.order_number}” deleted.")
        return super().form_valid(form)
    class SupportView(LoginRequiredMixin, TemplateView):
            template_name = "crud/support.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tickets = SupportTicket.objects.all()
        context["tickets"] = tickets
        context["all_count"] = tickets.count()
        context["needs_attention_count"] = tickets.filter(status="needs_attention").count()
        context["in_progress_count"] = tickets.filter(status="in_progress").count()
        context["closed_count"] = tickets.filter(status="closed").count()
        context["active_tab"] = self.request.GET.get("tab", "help")

        context["return_help_topics"] = [
            "I have received wrong return", "I have received damaged return",
            "I have not received my Return/RTO shipment", "Item/s are missing in my return",
            "I have received a wrong barcoded package in RTO", "I have received used product as return",
            "Return/RTO product not received but marked delivered - Need Proof of Delivery",
            "I have an issue with Exchange order",
            "Return/RTO Delivery Issue - False Attempt by Logistic Partner",
            "I am not able to generate invoice for exchange order",
            "I have received an RTO in a non-barcoded package",
            "I am unable to raise Wrong Return / RTO claims",
            "Order return shipping charge fee issue",
            "When will I receive my wrong return related compensation",
            "I want to stop using the Wrong/Defective Returns Feature",
            "My order has been marked as Returnless Refund. What is Returnless Refund?",
            "Other Returns/RTO and Exchange related issue",
        ]
        context["help_categories"] = [
            "Cataloging & Pricing", "Orders & Delivery", "Payments", "Inventory",
            "Account", "Advertisements & Promotions", "Instant Cash", "Others",
        ]
        return context
# --- Inventory: full CRUD (tables2 + django-filter + crispy form + messages) ---
class InventoryListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Inventory
    table_class = InventoryTable
    filterset_class = InventoryFilter
    template_name = "crud/inventory_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        qs = super().get_queryset().select_related("product", "warehouse")
        allowed_sorts = {"-quantity", "quantity", "-updated", "updated"}
        sort = self.request.GET.get("sort")
        if sort in allowed_sorts:
            qs = qs.order_by(sort)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_inventory = Inventory.objects.all()
        context["total_items"] = all_inventory.count()
        context["in_stock_count"] = all_inventory.filter(status="in_stock").count()
        context["low_stock_count"] = all_inventory.filter(status="low_stock").count()
        context["out_of_stock_count"] = all_inventory.filter(status="out_of_stock").count()
        context["discontinued_count"] = all_inventory.filter(status="discontinued").count()

        all_products = Product.objects.all()
        context["active_products_count"] = all_products.filter(status="active").count()
        context["draft_products_count"] = all_products.filter(status="draft").count()
        context["inactive_products_count"] = all_products.filter(status="inactive").count()

        context["current_sort"] = self.request.GET.get("sort", "-quantity")
        return context
class InventoryCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Inventory
    form_class = InventoryForm
    template_name = "crud/inventory_form.html"
    success_url = reverse_lazy("crud:inventory_list")

    def form_valid(self, form):
        messages.success(self.request, f"Inventory created for “{form.instance.product.name}”.")
        return super().form_valid(form)


class InventoryUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Inventory
    form_class = InventoryForm
    template_name = "crud/inventory_form.html"
    success_url = reverse_lazy("crud:inventory_list")

    def form_valid(self, form):
        messages.success(self.request, f"Inventory for “{form.instance.product.name}” updated.")
        return super().form_valid(form)


class InventoryDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Inventory
    template_name = "crud/inventory_confirm_delete.html"
    success_url = reverse_lazy("crud:inventory_list")

    def form_valid(self, form):
        messages.success(self.request, f"Inventory for “{self.object.product.name}” deleted.")
        return super().form_valid(form)



class InventoryBulkStockUpdateView(LoginRequiredMixin, View):
    """Handles the Bulk Stock Update modal (Step 2: Upload)."""
    template_name = "crud/inventory_bulk_stock_form.html"

    def get(self, request, *args, **kwargs):
        form = InventoryBulkStockUploadForm()
        html = render_to_string(self.template_name, {"form": form}, request=request)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"html": html})
        return HttpResponse(html)

    def post(self, request, *args, **kwargs):
        form = InventoryBulkStockUploadForm(request.POST, request.FILES)
        if not form.is_valid():
            html = render_to_string(self.template_name, {"form": form}, request=request)
            return JsonResponse({"success": False, "html": html})

        uploaded_file = request.FILES["file"]
        updated_count = 0
        skipped = 0

        try:
            decoded = uploaded_file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            for row in reader:
                inv_id = row.get("Inventory ID") or row.get("id")
                qty = row.get("Quantity") or row.get("quantity")
                if not inv_id or qty in (None, ""):
                    skipped += 1
                    continue
                try:
                    inv = Inventory.objects.get(pk=int(inv_id))
                    inv.quantity = int(qty)
                    inv.save()
                    updated_count += 1
                except (Inventory.DoesNotExist, ValueError):
                    skipped += 1
        except Exception as e:
            form.add_error("file", f"File could not be processed: {e}")
            html = render_to_string(self.template_name, {"form": form}, request=request)
            return JsonResponse({"success": False, "html": html})

        msg = f"Bulk stock update complete — {updated_count} item(s) updated."
        if skipped:
            msg += f" {skipped} row(s) skipped."
        messages.success(request, msg)
        return JsonResponse({"success": True})


class InventoryBulkStockDownloadView(LoginRequiredMixin, View):
    """Handles Step 1: Download current stock as CSV."""

    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="inventory_stock.csv"'
        writer = csv.writer(response)
        writer.writerow(["Inventory ID", "Product", "Warehouse", "Quantity", "Reorder Level", "Status"])
        for inv in Inventory.objects.select_related("product", "warehouse").all():
            writer.writerow([
                inv.id,
                inv.product.name,
                inv.warehouse.name if inv.warehouse else "",
                inv.quantity,
                inv.reorder_level,
                inv.get_status_display(),
            ])
        return response

# --- Catalog Uploads: full CRUD (tables2 + django-filter + crispy form + messages) ---
class CatalogUploadListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = CatalogUpload
    table_class = CatalogUploadTable
    filterset_class = CatalogUploadFilter
    template_name = "crud/catalogupload_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        qs = super().get_queryset()
        upload_type = self.request.GET.get("upload_type", "bulk")
        return qs.filter(upload_type=upload_type)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_uploads = CatalogUpload.objects.all()
        context["total_uploads"] = all_uploads.count()
        context["bulk_count"] = all_uploads.filter(upload_type="bulk").count()
        context["single_count"] = all_uploads.filter(upload_type="single").count()
        context["active_upload_type"] = self.request.GET.get("upload_type", "bulk")

        current_type_qs = all_uploads.filter(upload_type=context["active_upload_type"])
        context["action_required_count"] = current_type_qs.filter(status="pending").count()
        context["qc_progress_count"] = current_type_qs.filter(status="processing").count()
        context["qc_error_count"] = current_type_qs.filter(status="failed").count()
        context["qc_pass_count"] = current_type_qs.filter(status="completed").count()
        context["active_status_tab"] = self.request.GET.get("status", "")
        return context


class CatalogUploadCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = CatalogUpload
    form_class = CatalogUploadForm
    template_name = "crud/catalogupload_form.html"
    success_url = reverse_lazy("crud:catalogupload_list")

    def form_valid(self, form):
        messages.success(self.request, f"Catalog upload “{form.instance.file_name}” created.")
        return super().form_valid(form)


class CatalogUploadUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = CatalogUpload
    form_class = CatalogUploadForm
    template_name = "crud/catalogupload_form.html"
    success_url = reverse_lazy("crud:catalogupload_list")

    def form_valid(self, form):
        messages.success(self.request, f"Catalog upload “{form.instance.file_name}” updated.")
        return super().form_valid(form)


class CatalogUploadDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = CatalogUpload
    template_name = "crud/catalogupload_confirm_delete.html"
    success_url = reverse_lazy("crud:catalogupload_list")

    def form_valid(self, form):
        messages.success(self.request, f"Catalog upload “{self.object.file_name}” deleted.")
        return super().form_valid(form)


# --- Image Bulk Uploads: full CRUD (tables2 + django-filter + crispy form + messages) ---
class ImageBulkUploadListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = ImageBulkUpload
    table_class = ImageBulkUploadTable
    filterset_class = ImageBulkUploadFilter
    template_name = "crud/imagebulkupload_list.html"
    table_pagination = {"per_page": 10}


class ImageBulkUploadCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = ImageBulkUpload
    form_class = ImageBulkUploadForm
    template_name = "crud/imagebulkupload_form.html"
    success_url = reverse_lazy("crud:imagebulkupload_list")

    def form_valid(self, form):
        messages.success(self.request, f"Image bulk upload “{form.instance.name}” created.")
        return super().form_valid(form)


class ImageBulkUploadUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = ImageBulkUpload
    form_class = ImageBulkUploadForm
    template_name = "crud/imagebulkupload_form.html"
    success_url = reverse_lazy("crud:imagebulkupload_list")

    def form_valid(self, form):
        messages.success(self.request, f"Image bulk upload “{form.instance.name}” updated.")
        return super().form_valid(form)


class ImageBulkUploadDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = ImageBulkUpload
    template_name = "crud/imagebulkupload_confirm_delete.html"
    success_url = reverse_lazy("crud:imagebulkupload_list")

    def form_valid(self, form):
        messages.success(self.request, f"Image bulk upload “{self.object.name}” deleted.")
        return super().form_valid(form)


# --- Quality: full CRUD (tables2 + django-filter + crispy form + messages) ---
class QualityListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Quality
    table_class = QualityTable
    filterset_class = QualityFilter
    template_name = "crud/quality_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return super().get_queryset().select_related("product")


class QualityCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Quality
    form_class = QualityForm
    template_name = "crud/quality_form.html"
    success_url = reverse_lazy("crud:quality_list")

    def form_valid(self, form):
        messages.success(self.request, f"Quality check for “{form.instance.product.name}” created.")
        return super().form_valid(form)


class QualityUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Quality
    form_class = QualityForm
    template_name = "crud/quality_form.html"
    success_url = reverse_lazy("crud:quality_list")

    def form_valid(self, form):
        messages.success(self.request, f"Quality check for “{form.instance.product.name}” updated.")
        return super().form_valid(form)


class QualityDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Quality
    template_name = "crud/quality_confirm_delete.html"
    success_url = reverse_lazy("crud:quality_list")

    def form_valid(self, form):
        messages.success(self.request, f"Quality check for “{self.object.product.name}” deleted.")
        return super().form_valid(form)
class QualityDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "crud/quality_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_quality = Quality.objects.select_related("product").all()

        context["total_checks"] = all_quality.count()
        context["blocking_soon"] = all_quality.filter(status="failed")
        context["action_pending"] = all_quality.filter(status__in=["pending", "rework"])
        context["fixed"] = all_quality.filter(status="passed")

        context["blocking_soon_count"] = context["blocking_soon"].count()
        context["action_pending_count"] = context["action_pending"].count()
        context["fixed_count"] = context["fixed"].count()

        total_defects = sum(q.defect_count for q in all_quality)
        avg_defects = round(total_defects / all_quality.count(), 1) if all_quality.count() else 0
        context["avg_defects"] = avg_defects
        context["has_score"] = all_quality.count() > 0

        context["active_tab"] = self.request.GET.get("tab", "action_pending")
        return context

class BlockedProductsView(LoginRequiredMixin, TemplateView):
    template_name = "crud/blocked_products.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_inventory = Inventory.objects.select_related("product")
        context["active_count"] = all_inventory.filter(status="in_stock").count()
        context["pending_count"] = all_inventory.filter(status="low_stock").count()
        context["blocked_count"] = all_inventory.filter(status="discontinued").count()
        context["paused_count"] = all_inventory.filter(status="out_of_stock").count()
        context["active_main_tab"] = self.request.GET.get("tab", "blocked")
        context["active_sub_tab"] = self.request.GET.get("subtab", "all")
        context["blocked_products"] = all_inventory.filter(status="discontinued")
        return context
    
# --- Products: full CRUD (tables2 + django-filter + crispy form + messages) ---
class ProductListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Product
    table_class = ProductTable
    filterset_class = ProductFilter
    template_name = "crud/product_list.html"
    table_pagination = {"per_page": 10}


class ProductCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = "crud/product_form.html"
    success_url = reverse_lazy("crud:product_list")

    def form_valid(self, form):
        messages.success(self.request, f"Product “{form.instance.name}” created.")
        return super().form_valid(form)


class ProductUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = "crud/product_form.html"
    success_url = reverse_lazy("crud:product_list")

    def form_valid(self, form):
        messages.success(self.request, f"Product “{form.instance.name}” updated.")
        return super().form_valid(form)


class ProductDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Product
    template_name = "crud/product_confirm_delete.html"
    success_url = reverse_lazy("crud:product_list")

    def form_valid(self, form):
        messages.success(self.request, f"Product “{self.object.name}” deleted.")
        return super().form_valid(form)


# --- Payments: full CRUD (tables2 + django-filter + crispy form + messages) ---
class PaymentListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Payment
    table_class = PaymentTable
    filterset_class = PaymentFilter
    template_name = "crud/payment_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return super().get_queryset().select_related("order")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        upcoming_cycles = PayoutCycle.objects.filter(status="upcoming").order_by("cycle_date")
        completed_cycles = PayoutCycle.objects.filter(status="completed").order_by("-cycle_date")

        context["upcoming_cycle"] = upcoming_cycles.first()
        context["upcoming_total"] = upcoming_cycles.aggregate(total=models.Sum("sales_returns"))["total"] or 0
        context["completed_cycle"] = completed_cycles.first()
        context["completed_cycles_list"] = completed_cycles[:3]
        context["completed_total"] = completed_cycles.aggregate(total=models.Sum("sales_returns"))["total"] or 0

        all_cycles = PayoutCycle.objects.all().order_by("cycle_date")
        context["chart_labels"] = [c.cycle_date.strftime("%d %b") for c in all_cycles]
        context["chart_payments"] = [float(c.net_amount) if c.status == "completed" else 0 for c in all_cycles]
        context["chart_outstanding"] = [float(c.net_amount) if c.status == "upcoming" else 0 for c in all_cycles]

        comp_qs = CompensationRecovery.objects.filter(record_type="compensation")
        rec_qs = CompensationRecovery.objects.filter(record_type="recovery")
        context["compensation_total"] = comp_qs.aggregate(total=models.Sum("amount"))["total"] or 0
        context["recovery_total"] = rec_qs.aggregate(total=models.Sum("amount"))["total"] or 0
        context["ads_cost_total"] = PayoutCycle.objects.aggregate(total=models.Sum("ads_cost"))["total"] or 0
        return context

class PaymentCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Payment
    form_class = PaymentForm
    template_name = "crud/payment_form.html"
    success_url = reverse_lazy("crud:payment_list")

    def form_valid(self, form):
        messages.success(self.request, f"Payment created for order “{form.instance.order.order_number}”.")
        return super().form_valid(form)


class PaymentUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Payment
    form_class = PaymentForm
    template_name = "crud/payment_form.html"
    success_url = reverse_lazy("crud:payment_list")

    def form_valid(self, form):
        messages.success(self.request, f"Payment for “{form.instance.order.order_number}” updated.")
        return super().form_valid(form)


class PaymentDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Payment
    template_name = "crud/payment_confirm_delete.html"
    success_url = reverse_lazy("crud:payment_list")

    def form_valid(self, form):
        messages.success(self.request, f"Payment for “{self.object.order.order_number}” deleted.")
        return super().form_valid(form)
from django.shortcuts import get_object_or_404
from django.http import Http404

PAYMENTS_FAQS = [
    {
        "slug": "not-received-payment",
        "q": "I have not received payments for my orders",
        "intro": "To check for payment settlement, please provide the Order ID / Sub Order ID below:",
        "show_order_lookup": True,
        "body": [
            "We prioritize timely payments for your orders, settling payments within 7 business days after delivery.",
        ],
        "steps": [
            "Go to the Payments Section",
            "You can use 'Search option' on the right top corner. You will be able to see the required details by using sub-order numbers or order numbers.",
        ],
        "note": "If the order is RTO/Cancelled then the payment will be '0' from Meesho's end.",
        "closing": "If you haven't received your payment even after 7 business days since delivery, kindly raise a ticket and our support team will assist in resolving the issue promptly.",
    },
    {"slug": "upcoming-payments", "q": "I want to know about my upcoming payments",
     "body": ["Upcoming payments are shown on the Payments dashboard along with the expected payout date and full breakdown of sales, deductions and benefits."]},
    {"slug": "gst-sales-report", "q": "I want GST or Sales report",
     "body": ["You can download your GST and Sales report anytime using the Download button on the Payments page."]},
    {"slug": "commission-invoice", "q": "I want to download the Commission Tax Invoice",
     "body": ["Commission Tax Invoices are generated for each completed payout cycle. If you can't locate one, please raise a ticket with the payout date."]},
    {"slug": "incorrect-invoice", "q": "I have received incorrect Invoice",
     "body": ["Please raise a ticket with the order details and our team will review and correct the invoice."]},
    {"slug": "change-bank-details", "q": "I want to change my bank details",
     "body": ["Bank details can be updated from your Account settings. Changes may take 2-3 business days to reflect in your payouts."]},
    {"slug": "tds-reimbursement", "q": "I want to file TDS reimbursement",
     "body": ["TDS reimbursement requests can be filed by raising a support ticket along with your TDS certificate attached."]},
    {"slug": "shipping-charges", "q": "I want to know about Shipping charges",
     "body": ["Shipping charges are deducted automatically based on order weight and destination, and are visible in your payout breakdown."]},
    {"slug": "settlement-calculation", "q": "I want to understand my settlement calculation",
     "body": ["Settlement = Sales Returns − Ads Cost − Program Cost + Program Benefits + Referral Earnings."]},
    {"slug": "referral-payments", "q": "I have not received payments for referrals",
     "body": ["Referral payments are processed monthly. Check the Referral Payments link on your Payments page for the latest status."]},
    {"slug": "lost-shipment-compensation", "q": "I have not received compensation for my lost shipment",
     "body": ["Compensation for lost shipments is credited automatically once the claim is approved by our team."]},
    {"slug": "returns-assurance", "q": "What is the Returns Assurance Program?",
     "body": ["It's a program that protects your margin against unexpected return costs by capping your return-related deductions."]},
    {"slug": "order-deduction", "q": "I want to know about the deduction for my order",
     "body": ["Deductions are listed transaction-wise in the payout details for each cycle, including ads cost, program cost and shipping."]},
    {"slug": "other-payment-issues", "q": "Other Payment related issues",
     "body": ["For any other payment issue not listed here, please raise a support ticket and our team will assist you."]},
]


class PaymentsHelpView(LoginRequiredMixin, TemplateView):
    template_name = "crud/payments_help.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["faqs"] = PAYMENTS_FAQS
        return context


class PaymentsHelpDetailView(LoginRequiredMixin, TemplateView):
    template_name = "crud/payments_help_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        slug = kwargs.get("slug")
        faq = next((f for f in PAYMENTS_FAQS if f["slug"] == slug), None)
        if not faq:
            raise Http404("FAQ not found")
        context["faq"] = faq
        context["other_faqs"] = [f for f in PAYMENTS_FAQS if f["slug"] != slug]
        return context
class PaymentsMyTicketsView(LoginRequiredMixin, TemplateView):
    template_name = "crud/payments_my_tickets.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tickets = SupportTicket.objects.all()
        active_tab = self.request.GET.get("tab", "all")
        context["active_tab"] = active_tab
        context["all_count"] = tickets.count()
        context["needs_attention_count"] = tickets.filter(status="needs_attention").count()
        context["in_progress_count"] = tickets.filter(status="in_progress").count()
        context["closed_count"] = tickets.filter(status="closed").count()

        if active_tab == "needs_attention":
            tickets = tickets.filter(status="needs_attention")
        elif active_tab == "in_progress":
            tickets = tickets.filter(status="in_progress")
        elif active_tab == "closed":
            tickets = tickets.filter(status="closed")

        context["tickets"] = tickets
        return context
# --- Warehouses: full CRUD (tables2 + django-filter + crispy form + messages) ---
class WarehouseListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Warehouse
    table_class = WarehouseTable
    filterset_class = WarehouseFilter
    template_name = "crud/warehouse_list.html"
    table_pagination = {"per_page": 10}


class WarehouseCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Warehouse
    form_class = WarehouseForm
    template_name = "crud/warehouse_form.html"
    success_url = reverse_lazy("crud:warehouse_list")

    def form_valid(self, form):
        messages.success(self.request, f"Warehouse “{form.instance.name}” created.")
        return super().form_valid(form)


class WarehouseUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Warehouse
    form_class = WarehouseForm
    template_name = "crud/warehouse_form.html"
    success_url = reverse_lazy("crud:warehouse_list")

    def form_valid(self, form):
        messages.success(self.request, f"Warehouse “{form.instance.name}” updated.")
        return super().form_valid(form)


class WarehouseDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Warehouse
    template_name = "crud/warehouse_confirm_delete.html"
    success_url = reverse_lazy("crud:warehouse_list")

    def form_valid(self, form):
        messages.success(self.request, f"Warehouse “{self.object.name}” deleted.")
        return super().form_valid(form)

class CallbackRequestCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = CallbackRequest
    form_class =  CallbackRequestForm
    template_name = "crud/callbackrequest_form.html"
    success_url = reverse_lazy("crud:warehouse_list")

    def form_valid(self, form):
        messages.success(self.request, "Your callback request has been submitted. Our team will reach out shortly.")
        return super().form_valid(form)
# --- Influencer Campaigns: full CRUD (tables2 + django-filter + crispy form + messages) ---
class InfluencerCampaignListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = InfluencerCampaign
    table_class = InfluencerCampaignTable
    filterset_class = InfluencerCampaignFilter
    template_name = "crud/influencercampaign_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return super().get_queryset().select_related("product")


class InfluencerCampaignCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = InfluencerCampaign
    form_class = InfluencerCampaignForm
    template_name = "crud/influencercampaign_form.html"
    success_url = reverse_lazy("crud:influencercampaign_list")

    def form_valid(self, form):
        messages.success(self.request, f"Campaign “{form.instance.campaign_name}” created.")
        return super().form_valid(form)


class InfluencerCampaignUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = InfluencerCampaign
    form_class = InfluencerCampaignForm
    template_name = "crud/influencercampaign_form.html"
    success_url = reverse_lazy("crud:influencercampaign_list")

    def form_valid(self, form):
        messages.success(self.request, f"Campaign “{form.instance.campaign_name}” updated.")
        return super().form_valid(form)


class InfluencerCampaignDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = InfluencerCampaign
    template_name = "crud/influencercampaign_confirm_delete.html"
    success_url = reverse_lazy("crud:influencercampaign_list")

    def form_valid(self, form):
        messages.success(self.request, f"Campaign “{self.object.campaign_name}” deleted.")
        return super().form_valid(form)


# --- Advertisements: full CRUD (tables2 + django-filter + crispy form + messages) ---
class AdvertisementListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Advertisement
    table_class = AdvertisementTable
    filterset_class = AdvertisementFilter
    template_name = "crud/advertisement_list.html"
    table_pagination = {"per_page": 10}

    def get_queryset(self):
        return super().get_queryset().select_related("product")


class AdvertisementCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Advertisement
    form_class = AdvertisementForm
    template_name = "crud/advertisement_form.html"
    success_url = reverse_lazy("crud:advertisement_list")

    def form_valid(self, form):
        messages.success(self.request, f"Advertisement “{form.instance.campaign_name}” created.")
        return super().form_valid(form)


class AdvertisementUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Advertisement
    form_class = AdvertisementForm
    template_name = "crud/advertisement_form.html"
    success_url = reverse_lazy("crud:advertisement_list")

    def form_valid(self, form):
        messages.success(self.request, f"Advertisement “{form.instance.campaign_name}” updated.")
        return super().form_valid(form)


class AdvertisementDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Advertisement
    template_name = "crud/advertisement_confirm_delete.html"
    success_url = reverse_lazy("crud:advertisement_list")

    def form_valid(self, form):
        messages.success(self.request, f"Advertisement “{self.object.campaign_name}” deleted.")
        return super().form_valid(form)
from django.views.generic import TemplateView


class BarcodeScanView(LoginRequiredMixin, TemplateView):
    template_name = "crud/barcode_scan.html"
from django.views.generic import TemplateView
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import redirect

class ReduceRTOReturnsView(LoginRequiredMixin, TemplateView):
    template_name = "crud/reduce_rto_returns.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["products"] = Product.objects.all()
        context["total_products"] = Product.objects.count()
        return context

    def post(self, request, *args, **kwargs):
        product_ids = request.POST.getlist("product_ids")
        if not product_ids:
            messages.warning(request, "No products selected.")
            return redirect("crud:reduce_rto_returns")

        applied_count = 0
        for pid in product_ids:
            prepaid_discount = request.POST.get(f"prepaid_discount_{pid}") or 0
            wdrp_discount = request.POST.get(f"wdrp_discount_{pid}") or 0
            try:
                product = Product.objects.get(pk=pid)
            except Product.DoesNotExist:
                continue

            discount_value = float(prepaid_discount) if float(prepaid_discount or 0) > 0 else float(wdrp_discount or 0)

            pricing, created = Pricing.objects.get_or_create(
                product=product,
                defaults={
                    "cost_price": product.price,
                    "selling_price": product.price,
                    "discount_percent": discount_value,
                    "status": "active",
                    "effective_date": timezone.now().date(),
                },
            )
            if not created:
                pricing.discount_percent = discount_value
                pricing.status = "active"
                pricing.save()
            applied_count += 1

        messages.success(request, f"Discount applied to {applied_count} product(s).")
        return redirect("crud:pricing_list")
class BulkCatalogCategoryView(LoginRequiredMixin, TemplateView):
    template_name = "crud/bulk_catalog_category.html"
class SupportView(LoginRequiredMixin, TemplateView):
    template_name = "crud/support.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        tickets = SupportTicket.objects.all()
        context["tickets"] = tickets
        context["all_count"] = tickets.count()
        context["needs_attention_count"] = tickets.filter(status="needs_attention").count()
        context["in_progress_count"] = tickets.filter(status="in_progress").count()
        context["closed_count"] = tickets.filter(status="closed").count()
        context["active_tab"] = self.request.GET.get("tab", "help")

        context["return_help_topics"] = [
            "I have received wrong return", "I have received damaged return",
            "I have not received my Return/RTO shipment", "Item/s are missing in my return",
            "I have received a wrong barcoded package in RTO", "I have received used product as return",
            "Return/RTO product not received but marked delivered - Need Proof of Delivery",
            "I have an issue with Exchange order",
            "Return/RTO Delivery Issue - False Attempt by Logistic Partner",
            "I am not able to generate invoice for exchange order",
            "I have received an RTO in a non-barcoded package",
            "I am unable to raise Wrong Return / RTO claims",
            "Order return shipping charge fee issue",
            "When will I receive my wrong return related compensation",
            "I want to stop using the Wrong/Defective Returns Feature",
            "My order has been marked as Returnless Refund. What is Returnless Refund?",
            "Other Returns/RTO and Exchange related issue",
        ]
        context["help_categories"] = [
            "Cataloging & Pricing", "Orders & Delivery", "Payments", "Inventory",
            "Account", "Advertisements & Promotions", "Instant Cash", "Others",
        ]
        return context
class ImageBulkUploadPageView(LoginRequiredMixin, TemplateView):
    template_name = "crud/imagebulkupload_upload.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["not_allowed_types"] = [
            "Watermark image", "Fake branded/1st copy", "Image with price",
            "Pixelated image", "Inverted image", "Blur/unclear image",
            "Incomplete image", "Stretched/shrunk image", "Image with props", "Image with text",
        ]
        return context
class InfluencerMarketingDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "crud/influencer_marketing_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["total_campaigns"] = InfluencerCampaign.objects.count()
        context["faqs"] = [
            ("Where can I access the Terms and Conditions of the Influencer Marketing Program?",
             "Please click here to view the detailed Terms and Conditions."),
            ("How can I opt in to the program?", "Go to Influencer Marketing and click Select Catalogs to opt in."),
            ("What is Influencer Marketing/Affiliate Marketing Program?",
             "A program where creators make videos on your products to drive orders."),
            ("How does the influencer marketing program work?",
             "Creators post content on social media and Meesho App; orders from that content are attributed to your catalogs."),
            ("Why should I participate in this program?", "It increases visibility and sales at no upfront cost."),
            ("Are there any charges for opt-in or content making?", "No, there are no charges for opt-in or content creation."),
            ("Can I set my own creator commission?", "Yes, you can set a commission between 4% and 20%."),
            ("What happens with non delivered orders coming via this program?",
             "There is no charge for non-delivered orders (return/RTO/cancellations). You pay only when the order is successfully delivered."),
            ("What are the social media channels via which my catalogs would be promoted?",
             "Your catalogs will be promoted by creators on YouTube, Instagram, Telegram and Facebook, and shown as reels on the Meesho App."),
            ("How can I opt out from the program?", "You can opt out anytime from the program in just one click."),
        ]
        return context
class InfluencerSelectCatalogsView(LoginRequiredMixin, TemplateView):
    template_name = "crud/influencer_select_catalogs.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["products"] = Product.objects.filter(status="active")
        return context
class AdvertisementDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "crud/advertisement_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_ads = Advertisement.objects.select_related("product").all()

        context["total_ad_spend"] = sum(a.spent for a in all_ads)
        context["total_clicks"] = sum(a.clicks for a in all_ads)
        context["total_impressions"] = sum(a.impressions for a in all_ads)
        context["total_budget"] = sum(a.budget for a in all_ads)

        context["all_count"] = all_ads.count()
        context["live_count"] = all_ads.filter(status="active").count()
        context["paused_count"] = all_ads.filter(status="paused").count()
        context["upcoming_count"] = all_ads.filter(status="draft").count()

        active_tab = self.request.GET.get("tab", "all")
        context["active_tab"] = active_tab
        if active_tab == "live":
            context["campaigns"] = all_ads.filter(status="active")
        elif active_tab == "paused":
            context["campaigns"] = all_ads.filter(status="paused")
        elif active_tab == "upcoming":
            context["campaigns"] = all_ads.filter(status="draft")
        else:
            context["campaigns"] = all_ads
        return context
# --- Promotions: full CRUD (tables2 + django-filter + crispy form + messages) ---
class PromotionListView(LoginRequiredMixin, SingleTableMixin, FilterView):
    model = Promotion
    table_class = PromotionTable
    filterset_class = PromotionFilter
    template_name = "crud/promotion_list.html"
    table_pagination = {"per_page": 10}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        all_promos = Promotion.objects.all()
        context["upcoming_count"] = all_promos.filter(status="upcoming").count()
        context["live_count"] = all_promos.filter(status="live").count()
        context["expired_count"] = all_promos.filter(status="expired").count()
        context["active_status_tab"] = self.request.GET.get("status_tab", "upcoming")
        return context


class PromotionCreateView(AjaxModalFormMixin, LoginRequiredMixin, CreateView):
    model = Promotion
    form_class = PromotionForm
    template_name = "crud/promotion_form.html"
    success_url = reverse_lazy("crud:promotion_list")

    def form_valid(self, form):
        messages.success(self.request, f"Promotion “{form.instance.event_name}” created.")
        return super().form_valid(form)


class PromotionUpdateView(AjaxModalFormMixin, LoginRequiredMixin, UpdateView):
    model = Promotion
    form_class = PromotionForm
    template_name = "crud/promotion_form.html"
    success_url = reverse_lazy("crud:promotion_list")

    def form_valid(self, form):
        messages.success(self.request, f"Promotion “{form.instance.event_name}” updated.")
        return super().form_valid(form)


class PromotionDeleteView(AjaxModalDeleteMixin, LoginRequiredMixin, DeleteView):
    model = Promotion
    template_name = "crud/promotion_confirm_delete.html"
    success_url = reverse_lazy("crud:promotion_list")

    def form_valid(self, form):
        messages.success(self.request, f"Promotion “{self.object.event_name}” deleted.")
        return super().form_valid(form)
class InstantCashView(LoginRequiredMixin, TemplateView):
    template_name = "crud/instant_cash.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["is_interested"] = self.request.session.get("instant_cash_interested", False)
        return context

    def post(self, request, *args, **kwargs):
        request.session["instant_cash_interested"] = True
        messages.success(request, "Thanks! You've been added to our waiting list.")
        return redirect("crud:instant_cash")
class InstantCashHelpView(LoginRequiredMixin, TemplateView):
    template_name = "crud/instant_cash_help.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["help_sections"] = [
            {
                "title": "Technical & App Issues",
                "items": ["App Crash / Loading / Timeout", "OTP / Login / Network Errors", "Buttons / Navigation / UI Errors"],
            },
            {"title": "Application and Eligibility", "items": []},
            {"title": "KYC and Selfie", "items": []},
            {"title": "Doc Signing and Verification and Auto-Debit Setup", "items": []},
            {"title": "Loan Offer & Terms Clarification", "items": []},
            {"title": "Disbursal & Account Credit", "items": []},
            {"title": "Others / General Queries", "items": []},
        ]
        return context
from datetime import timedelta
from django.db.models import Sum, Count


class BusinessDashboardView(LoginRequiredMixin, TemplateView):
    template_name = "crud/business_dashboard.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        period = self.request.GET.get("period", "last7")
        today = timezone.now().date()

        if period == "yesterday":
            start_date = end_date = today - timedelta(days=1)
        elif period == "last30":
            start_date, end_date = today - timedelta(days=29), today
        else:
            start_date, end_date = today - timedelta(days=6), today
            period = "last7"

        context["active_period"] = period
        context["date_range_label"] = f"{start_date.day} {start_date.strftime('%b %y')} - {end_date.day} {end_date.strftime('%b %y')}"
        orders_qs = Order.objects.filter(order_date__date__range=[start_date, end_date])
        total_orders = orders_qs.count()
        total_sales = orders_qs.aggregate(total=Sum("amount"))["total"] or 0

        metrics_qs = DailyMetric.objects.filter(date__range=[start_date, end_date])
        total_views = metrics_qs.aggregate(total=Sum("total_views"))["total"] or 0
        total_clicks = metrics_qs.aggregate(total=Sum("total_clicks"))["total"] or 0

        returns_qs = Return.objects.filter(order__in=orders_qs)
        return_percentage = round((returns_qs.count() / total_orders) * 100, 1) if total_orders else None
        conversion_rate = round((total_orders / total_clicks) * 100, 1) if total_clicks else 0

        context["total_views"] = total_views
        context["total_clicks"] = total_clicks
        context["total_orders"] = total_orders
        context["total_sales"] = total_sales
        context["conversion_rate"] = conversion_rate
        context["return_percentage"] = return_percentage

        chart_labels = []
        chart_orders = []
        chart_sales = []
        num_days = (end_date - start_date).days + 1
        for i in range(num_days):
            day = start_date + timedelta(days=i)
            day_orders = Order.objects.filter(order_date__date=day)
            chart_labels.append(day.strftime("%d %b"))
            chart_orders.append(day_orders.count())
            chart_sales.append(float(day_orders.aggregate(total=Sum("amount"))["total"] or 0))
        context["chart_labels"] = chart_labels
        context["chart_orders"] = chart_orders
        context["chart_sales"] = chart_sales
        context["has_trend"] = total_orders > 0

        context["top_products"] = (
            orders_qs.values("product__name")
            .annotate(order_count=Count("id"), sales_total=Sum("amount"))
            .order_by("-order_count")[:5]
        )
        return context


class StoreReportsView(LoginRequiredMixin, TemplateView):
    template_name = "crud/store_reports.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["orders"] = Order.objects.select_related("product").order_by("-order_date")[:20]
        context["metrics"] = DailyMetric.objects.all().order_by("-date")[:20]
        return context

# --- Sales Report ---
class SalesReportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        orders = Order.objects.select_related("product").order_by("-order_date")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        headers = ["Order Number", "Customer", "Product", "Quantity", "Amount", "Status", "Order Date"]
        ws.append(headers)
        for o in orders:
            ws.append([
                o.order_number, o.customer_name, o.product.name if o.product else "",
                o.quantity, float(o.amount), o.get_status_display(),
                o.order_date.strftime("%Y-%m-%d %H:%M"),
            ])
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="sales_report.xlsx"'
        wb.save(response)
        return response


class SalesReportPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        orders = Order.objects.select_related("product").order_by("-order_date")
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="sales_report.pdf"'
        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = [Paragraph("Sales Report", styles["Title"]), Spacer(1, 12)]

        data = [["Order #", "Customer", "Product", "Qty", "Amount", "Status", "Date"]]
        for o in orders:
            data.append([
                o.order_number, o.customer_name, o.product.name if o.product else "",
                str(o.quantity), f"${o.amount}", o.get_status_display(),
                o.order_date.strftime("%Y-%m-%d"),
            ])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)
        doc.build(elements)
        return response


# --- Visitors Report ---
class VisitorsReportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        metrics = DailyMetric.objects.all().order_by("-date")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visitors Report"
        headers = ["Date", "Total Views", "Total Clicks"]
        ws.append(headers)
        for m in metrics:
            ws.append([m.date.strftime("%Y-%m-%d"), m.total_views, m.total_clicks])
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="visitors_report.xlsx"'
        wb.save(response)
        return response


class VisitorsReportPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        metrics = DailyMetric.objects.all().order_by("-date")
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="visitors_report.pdf"'
        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = [Paragraph("Online Store Visitors Report", styles["Title"]), Spacer(1, 12)]

        data = [["Date", "Total Views", "Total Clicks"]]
        for m in metrics:
            data.append([m.date.strftime("%Y-%m-%d"), str(m.total_views), str(m.total_clicks)])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(table)
        doc.build(elements)
        return response
# --- Add these imports at the top of crud/views.py (if not already present) ---
# from django.contrib.auth.forms import PasswordChangeForm
# from django.contrib.auth import update_session_auth_hash
# from django.urls import reverse
# from shop.models import SupplierProfile
# from .forms import (
#     WhatsAppSettingsForm, BankDetailsForm, TaxDetailsForm,
#     SupplierSignatureForm, EmailNotificationsForm,
# )


# --- Settings: append this view to crud/views.py ---

LEGAL_POLICIES = [
    {"slug": "additional-seller-policies", "title": "Additional Seller Policies"},
    {"slug": "anti-phishing-policy", "title": "Anti Phishing Policy"},
    {"slug": "branded-packaging-policy", "title": "Branded Packaging Policy"},
    {"slug": "fair-usage-policy", "title": "Fair Usage Policy"},
    {"slug": "intellectual-property-policy", "title": "Intellectual Property Policy"},
    {"slug": "logistics-policy", "title": "Logistics Policy"},
    {"slug": "no-pack-program-policy", "title": "No-Pack Program Policy"},
    {"slug": "privacy-policy", "title": "Privacy Policy"},
    {"slug": "prohibited-restricted-products", "title": "Prohibited and Restricted Products list"},
    {"slug": "return-claims-policy", "title": "Return Claims Policy"},
    {"slug": "seller-agreement", "title": "Seller Agreement"},
    {"slug": "seller-deactivation-policy", "title": "Seller Deactivation Policy"},
    {"slug": "seller-referral-policy", "title": "Seller Referral Policy"},
    {"slug": "courier-partner-preference", "title": "T&Cs - Courier Partner Preference"},
    {"slug": "terms-and-conditions", "title": "Terms and Conditions"},
    {"slug": "whistle-blower-policy", "title": "Whistle Blower Policy"},
]

# Placeholder body content per policy (demo purposes only).
LEGAL_POLICY_BODY = {
    "additional-seller-policies": [
        "This policy outlines the additional obligations sellers agree to when listing products on the platform.",
        "Sellers must ensure that pricing, product descriptions, and images accurately represent the item being sold.",
        "The platform reserves the right to review and remove listings that violate marketplace guidelines.",
    ],
    "anti-phishing-policy": [
        "The platform will never ask sellers to share OTPs, passwords, or banking credentials over calls or messages.",
        "Any communication claiming to be from the platform requesting sensitive information should be treated as suspicious and reported immediately.",
    ],
    "branded-packaging-policy": [
        "Sellers who opt for branded packaging must adhere to size, material, and labeling guidelines provided in the seller dashboard.",
        "Packaging costs for branded material are borne by the seller unless otherwise stated in a promotional agreement.",
    ],
    "fair-usage-policy": [
        "This policy governs acceptable use of seller tools, bulk upload systems, and promotional credits provided by the platform.",
        "Any attempt to manipulate ratings, reviews, or order volumes artificially may result in account suspension.",
    ],
    "intellectual-property-policy": [
        "Sellers must only list products they have the legal right to sell and must not infringe on any trademark, copyright, or patent.",
        "Repeated IP violations may lead to permanent removal from the platform.",
    ],
    "logistics-policy": [
        "This policy covers pickup timelines, packaging standards, and courier partner responsibilities for order fulfillment.",
        "Sellers are expected to hand over orders within the committed dispatch window to avoid penalties.",
    ],
    "no-pack-program-policy": [
        "Under the No-Pack Program, eligible sellers can ship products without additional packaging, subject to category and courier eligibility.",
        "Sellers must ensure product safety during transit even without secondary packaging.",
    ],
    "privacy-policy": [
        "This policy describes how seller and customer data is collected, used, and protected across the platform.",
        "Personal information is never shared with third parties except as required to fulfill orders or comply with law.",
    ],
    "prohibited-restricted-products": [
        "Certain product categories are prohibited or restricted from sale on the platform, including hazardous materials, counterfeit goods, and items requiring special licensing.",
        "A full category-wise list is available in the seller resource center.",
    ],
    "return-claims-policy": [
        "This policy explains how return claims are raised, reviewed, and resolved between sellers and the platform.",
        "Sellers can dispute a claim within the specified window by submitting supporting evidence.",
    ],
    "seller-agreement": [
        "This agreement governs the relationship between the seller and the platform, including commission structure, payment cycles, and dispute resolution.",
        "By continuing to sell on the platform, sellers agree to the terms outlined in this document.",
    ],
    "seller-deactivation-policy": [
        "Accounts may be deactivated for repeated policy violations, poor quality scores, or fraudulent activity.",
        "Sellers can appeal a deactivation decision by raising a support ticket within 15 days.",
    ],
    "seller-referral-policy": [
        "Sellers who refer new sellers to the platform may be eligible for referral bonuses, subject to the referred seller meeting minimum activity requirements.",
    ],
    "courier-partner-preference": [
        "Sellers can set preferred courier partners for order fulfillment, subject to serviceability in their region.",
        "The platform may override preferences in case of courier unavailability to ensure timely delivery.",
    ],
    "terms-and-conditions": [
        "These terms and conditions govern seller access and use of the supplier platform.",
        "Continued use of the platform constitutes acceptance of any updates made to these terms.",
    ],
    "whistle-blower-policy": [
        "This policy provides a confidential channel for sellers and employees to report unethical or fraudulent activity without fear of retaliation.",
        "All reports are reviewed by an independent compliance team.",
    ],
}


class SettingsView(LoginRequiredMixin, TemplateView):
    template_name = "crud/settings.html"

    def get_profile(self):
        profile, _ = SupplierProfile.objects.get_or_create(user=self.request.user)
        return profile

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.get_profile()

        context["active_tab"] = self.request.GET.get("tab", "change_password")
        context["password_form"] = kwargs.get("password_form") or PasswordChangeForm(user=self.request.user)
        context["whatsapp_form"] = kwargs.get("whatsapp_form") or WhatsAppSettingsForm(instance=profile)
        context["bank_form"] = kwargs.get("bank_form") or BankDetailsForm(instance=profile)
        context["tax_form"] = kwargs.get("tax_form") or TaxDetailsForm(instance=profile)
        context["signature_form"] = kwargs.get("signature_form") or SupplierSignatureForm(instance=profile)
        context["email_form"] = kwargs.get("email_form") or EmailNotificationsForm(instance=profile)

        # Legal and Policies: list + detail
        context["legal_policies"] = LEGAL_POLICIES
        doc_slug = self.request.GET.get("doc")
        if doc_slug:
            policy = next((p for p in LEGAL_POLICIES if p["slug"] == doc_slug), None)
            if policy:
                context["active_policy"] = policy
                context["active_policy_body"] = LEGAL_POLICY_BODY.get(doc_slug, [])
        return context

    def post(self, request, *args, **kwargs):
        profile = self.get_profile()
        form_type = request.POST.get("form_type")

        if form_type == "change_password":
            form = PasswordChangeForm(user=request.user, data=request.POST)
            if form.is_valid():
                user = form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Password changed successfully.")
                return redirect(f"{reverse('crud:settings')}?tab=change_password")
            return self.render_to_response(self.get_context_data(password_form=form, active_tab="change_password"))

        elif form_type == "whatsapp":
            form = WhatsAppSettingsForm(request.POST, instance=profile)
            if form.is_valid():
                form.save()
                messages.success(request, "WhatsApp notification settings updated.")
                return redirect(f"{reverse('crud:settings')}?tab=whatsapp")
            return self.render_to_response(self.get_context_data(whatsapp_form=form, active_tab="whatsapp"))

        elif form_type == "bank":
            form = BankDetailsForm(request.POST, instance=profile)
            if form.is_valid():
                form.save()
                messages.success(request, "Bank details updated.")
                return redirect(f"{reverse('crud:settings')}?tab=bank")
            return self.render_to_response(self.get_context_data(bank_form=form, active_tab="bank"))

        elif form_type == "tax":
            form = TaxDetailsForm(request.POST, instance=profile)
            if form.is_valid():
                form.save()
                messages.success(request, "Tax details updated.")
                return redirect(f"{reverse('crud:settings')}?tab=tax")
            return self.render_to_response(self.get_context_data(tax_form=form, active_tab="tax"))

        elif form_type == "signature":
            form = SupplierSignatureForm(request.POST, request.FILES, instance=profile)
            if form.is_valid():
                form.save()
                messages.success(request, "Supplier signature updated.")
                return redirect(f"{reverse('crud:settings')}?tab=signature")
            return self.render_to_response(self.get_context_data(signature_form=form, active_tab="signature"))

        elif form_type == "email":
            form = EmailNotificationsForm(request.POST, instance=profile)
            if form.is_valid():
                form.save()
                messages.success(request, "Email notification settings updated.")
                return redirect(f"{reverse('crud:settings')}?tab=email")
            return self.render_to_response(self.get_context_data(email_form=form, active_tab="email"))

        messages.error(request, "Unknown form submitted.")
        return redirect("crud:settings")